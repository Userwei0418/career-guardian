"""Office document parser (Word/Excel)

Uses python-docx for Word documents and openpyxl for Excel files.
"""

from __future__ import annotations

from io import BytesIO

from app.services.parsers.base import DocumentParser, ParseResult


class OfficeParser(DocumentParser):
    """Microsoft Office document parser"""

    SUPPORTED_EXTENSIONS = {"docx", "doc", "xlsx", "xls"}

    def supports(self, filename: str) -> bool:
        return self._extension(filename) in self.SUPPORTED_EXTENSIONS

    def parse(self, file_bytes: bytes, filename: str) -> ParseResult:
        ext = self._extension(filename)
        if ext in ("docx", "doc"):
            return self._parse_word(file_bytes, filename)
        if ext in ("xlsx", "xls"):
            return self._parse_excel(file_bytes, filename)
        return ParseResult(
            raw_text="", page_count=0, parse_status="failed", parse_mode="office",
            parse_error=f"Unsupported office format: {ext}",
        )

    def _parse_word(self, file_bytes: bytes, filename: str) -> ParseResult:
        try:
            from docx import Document
            doc = Document(BytesIO(file_bytes))
            paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
            raw_text = "\n".join(paragraphs)
            return ParseResult(
                raw_text=raw_text, page_count=len(paragraphs), parse_status="completed", parse_mode="office",
                parse_notice=f"Extracted {len(paragraphs)} paragraphs from Word document.",
                metadata={"paragraph_count": len(paragraphs)},
            )
        except Exception as exc:
            return ParseResult(
                raw_text="", page_count=0, parse_status="failed", parse_mode="office",
                parse_error=f"Word parsing failed: {exc}",
            )

    def _parse_excel(self, file_bytes: bytes, filename: str) -> ParseResult:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            all_text = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        all_text.append(row_text.strip())
            raw_text = "\n".join(all_text)
            return ParseResult(
                raw_text=raw_text, page_count=len(wb.sheetnames), parse_status="completed", parse_mode="office",
                parse_notice=f"Extracted {len(all_text)} rows from {len(wb.sheetnames)} sheets.",
                metadata={"sheet_count": len(wb.sheetnames), "row_count": len(all_text)},
            )
        except Exception as exc:
            return ParseResult(
                raw_text="", page_count=0, parse_status="failed", parse_mode="office",
                parse_error=f"Excel parsing failed: {exc}",
            )